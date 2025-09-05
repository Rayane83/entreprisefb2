from sqlalchemy.orm import Session
from typing import List, Dict, Any, Optional
import re
import tempfile
import os
from datetime import datetime
import logging

from models import DotationReport, DotationRow

logger = logging.getLogger(__name__)

def calculate_dotation_totals(db: Session, report_id: str):
    """Calculer les totaux d'un rapport de dotation."""
    try:
        report = db.query(DotationReport).filter(DotationReport.id == report_id).first()
        if not report:
            return
        
        rows = db.query(DotationRow).filter(DotationRow.report_id == report_id).all()
        
        total_ca = sum(row.ca_total for row in rows)
        total_salaires = sum(row.salaire for row in rows)
        total_primes = sum(row.prime for row in rows)
        total_employees = len(rows)
        
        report.total_ca = total_ca
        report.total_salaires = total_salaires
        report.total_primes = total_primes
        report.total_employees = total_employees
        
        db.commit()
        
    except Exception as e:
        logger.error(f"Erreur lors du calcul des totaux: {e}")
        raise

def parse_bulk_dotation_data(data: str, format_type: str = "auto") -> List[Dict[str, Any]]:
    """Parser les données en lot pour l'import de dotations."""
    try:
        if not data or not data.strip():
            return []
        
        lines = data.strip().split('\n')
        parsed_rows = []
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
            
            # Détection automatique du séparateur
            if ';' in line:
                parts = line.split(';')
            elif '\t' in line:
                parts = line.split('\t')
            elif ',' in line:
                parts = line.split(',')
            else:
                parts = line.split()
            
            # Nettoyer les parties
            parts = [part.strip() for part in parts]
            
            if len(parts) >= 4:
                try:
                    nom = parts[0]
                    run = float(parts[1]) if parts[1] else 0.0
                    facture = float(parts[2]) if parts[2] else 0.0
                    vente = float(parts[3]) if parts[3] else 0.0
                    
                    # Calculs automatiques
                    ca_total = run + facture + vente
                    salaire = round(ca_total * 0.35)  # 35% du CA
                    prime = round(ca_total * 0.08)    # 8% du CA
                    
                    parsed_rows.append({
                        "nom": nom,
                        "grade": "À définir",
                        "run": run,
                        "facture": facture,
                        "vente": vente,
                        "ca_total": ca_total,
                        "salaire": salaire,
                        "prime": prime
                    })
                    
                except (ValueError, IndexError) as e:
                    logger.warning(f"Ligne ignorée (erreur de format): {line} - {e}")
                    continue
        
        return parsed_rows
        
    except Exception as e:
        logger.error(f"Erreur lors du parsing des données: {e}")
        return []

async def export_dotation_pdf(report: DotationReport, filters: Optional[Dict] = None) -> str:
    """Générer un export PDF d'un rapport de dotation."""
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        
        # Créer un fichier temporaire
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
        temp_path = temp_file.name
        temp_file.close()
        
        # Créer le document PDF
        doc = SimpleDocTemplate(temp_path, pagesize=landscape(A4))
        story = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=20,
            alignment=1  # Centré
        )
        
        # Titre
        title = Paragraph(f"Rapport de Dotation - {report.title}", title_style)
        story.append(title)
        
        # Informations générales
        info_data = [
            ["Période:", report.period or "Non spécifiée"],
            ["Nombre d'employés:", str(report.total_employees)],
            ["CA Total:", f"{report.total_ca:,.2f}€"],
            ["Salaires Total:", f"{report.total_salaires:,.2f}€"],
            ["Primes Total:", f"{report.total_primes:,.2f}€"],
            ["Date de création:", report.created_at.strftime("%d/%m/%Y %H:%M")]
        ]
        
        info_table = Table(info_data, colWidths=[4*cm, 6*cm])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.grey),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('BACKGROUND', (1, 0), (1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(info_table)
        story.append(Spacer(1, 20))
        
        # Tableau des employés
        if report.rows:
            headers = ["Nom", "Grade", "RUN", "FACTURE", "VENTE", "CA TOTAL", "Salaire", "Prime"]
            table_data = [headers]
            
            for row in report.rows:
                table_data.append([
                    row.employee_name,
                    row.grade or "N/A",
                    f"{row.run:,.0f}€",
                    f"{row.facture:,.0f}€",
                    f"{row.vente:,.0f}€",
                    f"{row.ca_total:,.0f}€",
                    f"{row.salaire:,.0f}€",
                    f"{row.prime:,.0f}€"
                ])
            
            employees_table = Table(table_data, colWidths=[3*cm, 2*cm, 2*cm, 2*cm, 2*cm, 2.5*cm, 2.5*cm, 2.5*cm])
            employees_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(employees_table)
        
        # Générer le PDF
        doc.build(story)
        
        logger.info(f"PDF généré: {temp_path}")
        return temp_path
        
    except Exception as e:
        logger.error(f"Erreur lors de la génération du PDF: {e}")
        raise

async def export_dotation_excel(report: DotationReport, filters: Optional[Dict] = None) -> str:
    """Générer un export Excel d'un rapport de dotation."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        # Créer un fichier temporaire
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_path = temp_file.name
        temp_file.close()
        
        # Créer le workbook
        wb = Workbook()
        
        # Feuille principale - Résumé
        ws_summary = wb.active
        ws_summary.title = "Résumé"
        
        # Styles
        title_font = Font(bold=True, size=16)
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Titre
        ws_summary['A1'] = f"Rapport de Dotation - {report.title}"
        ws_summary['A1'].font = title_font
        ws_summary.merge_cells('A1:F1')
        
        # Informations générales
        info_data = [
            ("Période", report.period or "Non spécifiée"),
            ("Nombre d'employés", report.total_employees),
            ("CA Total", f"{report.total_ca:,.2f}€"),
            ("Salaires Total", f"{report.total_salaires:,.2f}€"),
            ("Primes Total", f"{report.total_primes:,.2f}€"),
            ("Date de création", report.created_at.strftime("%d/%m/%Y %H:%M")),
            ("Statut", report.status.value)
        ]
        
        for i, (label, value) in enumerate(info_data, 3):
            ws_summary[f'A{i}'] = label
            ws_summary[f'B{i}'] = value
            ws_summary[f'A{i}'].font = header_font
        
        # Feuille détails employés
        if report.rows:
            ws_details = wb.create_sheet(title="Détails Employés")
            
            # Headers
            headers = ["Nom", "Grade", "RUN", "FACTURE", "VENTE", "CA TOTAL", "Salaire", "Prime", "Date"]
            for col, header in enumerate(headers, 1):
                cell = ws_details.cell(row=1, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Données
            for row_idx, row in enumerate(report.rows, 2):
                ws_details.cell(row=row_idx, column=1, value=row.employee_name)
                ws_details.cell(row=row_idx, column=2, value=row.grade or "N/A")
                ws_details.cell(row=row_idx, column=3, value=row.run)
                ws_details.cell(row=row_idx, column=4, value=row.facture)
                ws_details.cell(row=row_idx, column=5, value=row.vente)
                ws_details.cell(row=row_idx, column=6, value=row.ca_total)
                ws_details.cell(row=row_idx, column=7, value=row.salaire)
                ws_details.cell(row=row_idx, column=8, value=row.prime)
                ws_details.cell(row=row_idx, column=9, value=row.created_at.strftime("%d/%m/%Y"))
            
            # Ajuster la largeur des colonnes
            for column in ws_details.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_details.column_dimensions[column_letter].width = adjusted_width
        
        # Sauvegarder
        wb.save(temp_path)
        
        logger.info(f"Excel généré: {temp_path}")
        return temp_path
        
    except Exception as e:
        logger.error(f"Erreur lors de la génération du fichier Excel: {e}")
        raise